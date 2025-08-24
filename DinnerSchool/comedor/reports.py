import pandas as pd
from datetime import date, datetime
from django.db.models import Sum, Q, Count
from django.db import models
from comedor.models import Credito, CreditoDiario, Pedido
from core.choices import NIVELEDUCATIVO, GRADO, GRUPO
from core.herramientas import getChoiceLabel

import os
from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class ReporteGastosDiarios:
    """Clase para generar reportes de gastos diarios en Excel."""
    
    def __init__(self, fecha_reporte=None):
        """
        Inicializa el reporte con una fecha específica.
        Si no se proporciona fecha, usa la fecha actual.
        """
        self.fecha_reporte = fecha_reporte or date.today()
    
    def aplicar_estilos_excel(self, workbook):
        """Aplica estilos profesionales a todas las hojas del Excel."""
        
        # Definir colores del tema (puedes cambiar estos colores por los de tu front)
        colores = {
            'header_bg': 'dc2626',      # Azul principal
            'header_text': 'FFFFFFFF',     # Blanco
            'row_even': 'FFF8F9FA',       # Gris muy claro
            'row_odd': 'FFFFFFFF',        # Blanco
            'positive': 'FF28A745',       # Verde para valores positivos
            'negative': 'FFDC3545',       # Rojo para valores negativos
            'neutral': 'FF6C757D',        # Gris para neutro
            'border': 'FFD1D3E2'         # Gris para bordes
        }
        
        # Estilos base
        font_header = Font(name='Arial', size=12, bold=True, color=colores['header_text'])
        font_data = Font(name='Arial', size=10)
        font_title = Font(name='Arial', size=14, bold=True, color=colores['header_bg'])
        
        fill_header = PatternFill(start_color=colores['header_bg'], end_color=colores['header_bg'], fill_type='solid')
        fill_even = PatternFill(start_color=colores['row_even'], end_color=colores['row_even'], fill_type='solid')
        
        alignment_center = Alignment(horizontal='center', vertical='center')
        alignment_left = Alignment(horizontal='left', vertical='center')
        alignment_right = Alignment(horizontal='right', vertical='center')
        
        border_thin = Border(
            left=Side(style='thin', color=colores['border']),
            right=Side(style='thin', color=colores['border']),
            top=Side(style='thin', color=colores['border']),
            bottom=Side(style='thin', color=colores['border'])
        )
        
        # Aplicar estilos a cada hoja
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Agregar título de la hoja
            worksheet.insert_rows(1)
            worksheet['A1'] = f"REPORTE DINNERSCHOOL - {sheet_name}"
            worksheet['A1'].font = font_title
            worksheet['A1'].alignment = alignment_center
            worksheet.merge_cells('A1:J1')  # Ajustar según el número de columnas
            
            # Agregar fecha del reporte
            worksheet.insert_rows(2)
            worksheet['A2'] = f"Fecha del Reporte: {self.fecha_reporte.strftime('%d/%m/%Y')}"
            worksheet['A2'].font = Font(name='Arial', size=10, italic=True)
            worksheet['A2'].alignment = alignment_left
            
            # Agregar línea en blanco
            worksheet.insert_rows(3)
            
            # Los headers ahora están en la fila 4
            if worksheet.max_row > 3:
                # Aplicar estilo a headers
                for cell in worksheet[4]:
                    if cell.value:
                        cell.font = font_header
                        cell.fill = fill_header
                        cell.alignment = alignment_center
                        cell.border = border_thin
                
                # Aplicar estilos a datos
                for row_num in range(5, worksheet.max_row + 1):
                    for col_num in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        
                        # Alternar colores de filas
                        if (row_num - 5) % 2 == 0:
                            cell.fill = fill_even
                        
                        cell.font = font_data
                        cell.border = border_thin
                        
                        # Alineación según el tipo de dato
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = alignment_right
                            # Formato de moneda para columnas de dinero
                            if any(keyword in str(worksheet.cell(row=4, column=col_num).value).lower() 
                                  for keyword in ['monto', 'total', 'credito', 'deuda', 'balance']):
                                cell.number_format = '$#,##0.00'
                        elif isinstance(cell.value, date):
                            cell.alignment = alignment_center
                            cell.number_format = 'DD/MM/YYYY'
                        else:
                            cell.alignment = alignment_left
                        
                        # Colores especiales para estados
                        cell_value_str = str(cell.value).lower()
                        if 'positivo' in cell_value_str or 'crédito a favor' in cell_value_str:
                            cell.font = Font(name='Arial', size=10, color=colores['positive'], bold=True)
                        elif 'negativo' in cell_value_str or 'debe pagar' in cell_value_str:
                            cell.font = Font(name='Arial', size=10, color=colores['negative'], bold=True)
                        elif 'neutro' in cell_value_str or 'saldado' in cell_value_str:
                            cell.font = Font(name='Arial', size=10, color=colores['neutral'], bold=True)
            
            # Ajustar ancho de columnas
            self.ajustar_columnas(worksheet)
    
    def ajustar_columnas(self, worksheet):
        """Ajusta automáticamente el ancho de las columnas según el contenido."""
        column_widths = {}
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    column_letter = get_column_letter(cell.column)
                    content_length = len(str(cell.value))
                    
                    # Ajustes especiales para tipos de contenido
                    if isinstance(cell.value, date):
                        content_length = max(content_length, 12)  # Espacio para formato de fecha
                    elif 'ID' in str(cell.value):
                        content_length = max(content_length, 8)
                    elif any(keyword in str(cell.value).lower() for keyword in ['monto', 'total', 'balance']):
                        content_length = max(content_length, 15)  # Espacio para formato de moneda
                    elif 'estado' in str(cell.value).lower() or 'status' in str(cell.value).lower():
                        content_length = max(content_length, 20)
                    
                    if column_letter not in column_widths:
                        column_widths[column_letter] = content_length
                    else:
                        column_widths[column_letter] = max(column_widths[column_letter], content_length)
        
        # Aplicar anchos con límites mínimos y máximos
        for column_letter, width in column_widths.items():
            adjusted_width = min(max(width + 2, 10), 50)  # Mínimo 10, máximo 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def crear_hoja_con_estilo(self, writer, df, sheet_name, titulo_personalizado=None):
        """Crea una hoja con datos y aplica estilos automáticamente."""
        if df.empty:
            # Crear mensaje para hojas vacías
            df_mensaje = pd.DataFrame({
                'Información': [f'No hay datos disponibles para {sheet_name} en la fecha seleccionada'],
                'Fecha': [self.fecha_reporte],
                'Sugerencia': ['Verifique que existan registros para esta fecha']
            })
            df_mensaje.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    def obtener_datos_pedidos(self):
        """Obtiene los datos de pedidos del día."""
        pedidos = Pedido.objects.filter(fecha=self.fecha_reporte).select_related(
            'platillo', 'alumnoId', 'profesorId', 'nivelEducativo', 'encargadoId'
        )
        
        data_pedidos = []
        for pedido in pedidos:
            data_pedidos.append({
                'ID Pedido': pedido.id,
                'Platillo': pedido.platillo.nombre,
                'Ingredientes': pedido.ingredientePlatillo or pedido.platillo.ingredientes,
                'Nota': pedido.nota or '',
                'Alumno': str(pedido.alumnoId) if pedido.alumnoId else '',
                'Profesor': str(pedido.profesorId) if pedido.profesorId else '',
                'Nivel Educativo': f"{getChoiceLabel(NIVELEDUCATIVO,pedido.nivelEducativo.nivel)} - {getChoiceLabel(GRADO,pedido.nivelEducativo.grado)} - {getChoiceLabel(GRUPO,pedido.nivelEducativo.grupo)}" if pedido.nivelEducativo else '',
                'Total': float(pedido.total),
                'Estado': pedido.get_status_label(),
                'Turno': pedido.get_turno_label(),
                'Encargado': str(pedido.encargadoId) if pedido.encargadoId else '',
                'Fecha': pedido.fecha
            })
        
        return pd.DataFrame(data_pedidos)
    
    def obtener_datos_creditos(self):
        """Obtiene los datos de créditos diarios del día."""
        creditos_diarios = CreditoDiario.objects.filter(fecha=self.fecha_reporte).select_related(
            'tutorId', 'profesorId', 'pedido'
        )
        
        data_creditos = []
        for credito in creditos_diarios:
            # Determinar si es positivo o negativo
            status_credito = "Positivo" if credito.monto > 0 else "Negativo" if credito.monto < 0 else "Neutro"
            
            data_creditos.append({
                'ID Crédito Diario': credito.id,
                'Monto': float(credito.monto),
                'Status': status_credito,
                'Tutor': str(credito.tutorId) if credito.tutorId else '',
                'Profesor': str(credito.profesorId) if credito.profesorId else '',
                'Pedido Relacionado': credito.pedido.id if credito.pedido else '',
                'Fecha': credito.fecha,
                'Tipo': 'Crédito Diario'
            })
        
        return pd.DataFrame(data_creditos)
    
    def obtener_datos_creditos_generales(self):
        """Obtiene los datos de la tabla Credito del día."""
        creditos_generales = Credito.objects.filter(fecha=self.fecha_reporte).select_related(
            'tutorId', 'profesorId'
        )
        
        data_creditos_generales = []
        for credito in creditos_generales:
            # Determinar si es positivo o negativo
            status_credito = "Positivo" if credito.monto > 0 else "Negativo" if credito.monto < 0 else "Neutro"
            
            data_creditos_generales.append({
                'ID Crédito General': credito.id,
                'Monto': float(credito.monto),
                'Status': status_credito,
                'Tutor': str(credito.tutorId) if credito.tutorId else '',
                'Profesor': str(credito.profesorId) if credito.profesorId else '',
                'Fecha': credito.fecha,
                'Tipo': 'Crédito General'
            })
        
        return pd.DataFrame(data_creditos_generales)
    
    def obtener_resumen_gastos(self):
        """Calcula el resumen completo incluyendo créditos generales y diarios."""
        # Total de pedidos del día
        total_pedidos = Pedido.objects.filter(fecha=self.fecha_reporte).aggregate(
            total=Sum('total')
        )['total'] or 0
        
        # Créditos Diarios del día
        total_creditos_diarios = CreditoDiario.objects.filter(fecha=self.fecha_reporte).aggregate(
            total=Sum('monto')
        )['total'] or 0
        
        creditos_diarios_positivos = CreditoDiario.objects.filter(
            fecha=self.fecha_reporte, monto__gt=0
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        creditos_diarios_negativos = CreditoDiario.objects.filter(
            fecha=self.fecha_reporte, monto__lt=0
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        # Créditos Generales del día (tabla Credito)
        total_creditos_generales = Credito.objects.filter(fecha=self.fecha_reporte).aggregate(
            total=Sum('monto')
        )['total'] or 0
        
        creditos_generales_positivos = Credito.objects.filter(
            fecha=self.fecha_reporte, monto__gt=0
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        creditos_generales_negativos = Credito.objects.filter(
            fecha=self.fecha_reporte, monto__lt=0
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        # Totales combinados
        total_creditos_todos = total_creditos_diarios + total_creditos_generales
        total_positivos = creditos_diarios_positivos + creditos_generales_positivos
        total_negativos = creditos_diarios_negativos + creditos_generales_negativos
        
        # Balance final (lo que se debe pagar a la cafetería)
        deuda_cafeteria = abs(total_negativos)  # Convertir a positivo para mostrar la deuda
        credito_favor = total_positivos
        balance_neto = total_creditos_todos
        
        # Diferencia entre ingresos (créditos positivos) y gastos (pedidos + créditos negativos)
        gastos_totales = total_pedidos + deuda_cafeteria
        diferencia_final = credito_favor - gastos_totales
        
        resumen = {
            'Fecha del Reporte': self.fecha_reporte,
            # Pedidos
            'Total Pedidos (Ventas)': float(total_pedidos),
            # Créditos Diarios
            'Créditos Diarios Total': float(total_creditos_diarios),
            'Créditos Diarios Positivos': float(creditos_diarios_positivos),
            'Créditos Diarios Negativos': float(creditos_diarios_negativos),
            # Créditos Generales
            'Créditos Generales Total': float(total_creditos_generales),
            'Créditos Generales Positivos': float(creditos_generales_positivos),
            'Créditos Generales Negativos': float(creditos_generales_negativos),
            # Totales
            'Total Créditos Positivos': float(total_positivos),
            'Total Créditos Negativos': float(total_negativos),
            'DEUDA A CAFETERÍA': float(deuda_cafeteria),  # Lo que se debe pagar
            'Crédito a Favor': float(credito_favor),
            'Balance Neto': float(balance_neto),
            'Resultado Final': float(diferencia_final)
        }
        
        return resumen
    
    def obtener_reporte_deudas_creditos(self):
        """Genera un reporte específico de deudas por persona desde la tabla Credito."""
        creditos = Credito.objects.filter(fecha=self.fecha_reporte).select_related(
            'tutorId', 'profesorId'
        )
        
        # Diccionario para agrupar por persona
        deudas_por_persona = {}
        
        for credito in creditos:
            # Identificar la persona (tutor o profesor)
            persona = None
            tipo_persona = ""
            
            if credito.tutorId:
                persona = str(credito.tutorId)
                tipo_persona = "Tutor"
            elif credito.profesorId:
                persona = str(credito.profesorId)
                tipo_persona = "Profesor"
            else:
                persona = "Sin Asignar"
                tipo_persona = "N/A"
            
            # Crear clave única para la persona
            clave_persona = f"{persona} ({tipo_persona})"
            
            # Inicializar o actualizar los datos de la persona
            if clave_persona not in deudas_por_persona:
                deudas_por_persona[clave_persona] = {
                    'Nombre': persona,
                    'Tipo': tipo_persona,
                    'Total_Creditos': 0,
                    'Creditos_Positivos': 0,
                    'Creditos_Negativos': 0,
                    'Deuda_Total': 0,
                    'Credito_a_Favor': 0,
                    'Balance_Neto': 0,
                    'Cantidad_Movimientos': 0,
                    'Estado_Final': ''
                }
            
            # Acumular valores
            monto = float(credito.monto)
            deudas_por_persona[clave_persona]['Total_Creditos'] += monto
            deudas_por_persona[clave_persona]['Cantidad_Movimientos'] += 1
            
            if monto > 0:
                deudas_por_persona[clave_persona]['Creditos_Positivos'] += monto
                deudas_por_persona[clave_persona]['Credito_a_Favor'] += monto
            elif monto < 0:
                deudas_por_persona[clave_persona]['Creditos_Negativos'] += monto
                deudas_por_persona[clave_persona]['Deuda_Total'] += abs(monto)
        
        # Calcular balance neto y estado final para cada persona
        data_reporte = []
        for clave, datos in deudas_por_persona.items():
            balance_neto = datos['Total_Creditos']
            datos['Balance_Neto'] = balance_neto
            
            # Determinar estado final
            if balance_neto > 0:
                datos['Estado_Final'] = f"CRÉDITO A FAVOR: ${balance_neto:.2f}"
            elif balance_neto < 0:
                datos['Estado_Final'] = f"DEBE PAGAR: ${abs(balance_neto):.2f}"
            else:
                datos['Estado_Final'] = "SALDADO"
            
            data_reporte.append({
                'Persona': datos['Nombre'],
                'Tipo': datos['Tipo'],
                'Créditos Positivos': datos['Creditos_Positivos'],
                'Créditos Negativos': datos['Creditos_Negativos'],
                'Deuda Total': datos['Deuda_Total'],
                'Crédito a Favor': datos['Credito_a_Favor'],
                'Balance Neto': balance_neto,
                'Estado Final': datos['Estado_Final'],
                'Cantidad Movimientos': datos['Cantidad_Movimientos'],
                'Fecha': self.fecha_reporte
            })
        
        # Convertir a DataFrame y ordenar por balance (deudores primero)
        df_reporte = pd.DataFrame(data_reporte)
        if not df_reporte.empty:
            df_reporte = df_reporte.sort_values('Balance Neto', ascending=True)
        
        return df_reporte
    
    def obtener_reporte_deudas_detallado(self):
        """Genera un reporte detallado mostrando cada movimiento de crédito por persona."""
        creditos = Credito.objects.filter(fecha=self.fecha_reporte).select_related(
            'tutorId', 'profesorId'
        ).order_by('tutorId', 'profesorId', 'id')
        
        data_detallado = []
        for credito in creditos:
            # Identificar la persona
            persona = None
            tipo_persona = ""
            
            if credito.tutorId:
                persona = str(credito.tutorId)
                tipo_persona = "Tutor"
            elif credito.profesorId:
                persona = str(credito.profesorId)
                tipo_persona = "Profesor"
            else:
                persona = "Sin Asignar"
                tipo_persona = "N/A"
            
            # Determinar el estado del movimiento
            monto = float(credito.monto)
            if monto > 0:
                tipo_movimiento = "ABONO/CRÉDITO"
                descripcion = f"Crédito a favor"
            elif monto < 0:
                tipo_movimiento = "CARGO/DEUDA"
                descripcion = f"Deuda generada"
            else:
                tipo_movimiento = "NEUTRO"
                descripcion = "Sin efecto"
            
            data_detallado.append({
                'ID Crédito': credito.id,
                'Persona': persona,
                'Tipo Persona': tipo_persona,
                'Monto': monto,
                'Tipo Movimiento': tipo_movimiento,
                'Descripción': descripcion,
                'Fecha': credito.fecha
            })
        
        return pd.DataFrame(data_detallado)

    def generar_reporte_excel(self, nombre_archivo=None):
        """
        Genera el reporte completo en un archivo Excel con estilos profesionales.
        
        Args:
            nombre_archivo (str): Nombre del archivo. Si no se proporciona, 
                                se genera automáticamente.
        
        Returns:
            str: Ruta del archivo generado.
        """
        if not nombre_archivo:
            fecha_str = self.fecha_reporte.strftime("%Y%m%d")
            nombre_archivo = f"reporte_gastos_{fecha_str}.xlsx"
        
        # Crear directorio de reportes si no existe
        directorio_reportes = "reportes"
        if not os.path.exists(directorio_reportes):
            os.makedirs(directorio_reportes)
        
        ruta_archivo = os.path.join(directorio_reportes, nombre_archivo)
        
        # Obtener datos
        df_pedidos = self.obtener_datos_pedidos()
        df_creditos = self.obtener_datos_creditos()
        df_creditos_generales = self.obtener_datos_creditos_generales()
        df_deudas_resumen = self.obtener_reporte_deudas_creditos()
        df_deudas_detallado = self.obtener_reporte_deudas_detallado()
        resumen = self.obtener_resumen_gastos()
        
        # Crear DataFrame para el resumen con formato mejorado
        df_resumen = pd.DataFrame([resumen])
        
        # Escribir a Excel sin estilos primero
        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
            # Usar el método personalizado para crear hojas
            self.crear_hoja_con_estilo(writer, df_resumen, 'RESUMEN FINANCIERO')
            self.crear_hoja_con_estilo(writer, df_deudas_resumen, 'DEUDAS POR PERSONA')
            self.crear_hoja_con_estilo(writer, df_deudas_detallado, 'DETALLE CRÉDITOS')
            self.crear_hoja_con_estilo(writer, df_pedidos, 'Pedidos del Día')
            self.crear_hoja_con_estilo(writer, df_creditos, 'Créditos Diarios')
            self.crear_hoja_con_estilo(writer, df_creditos_generales, 'Créditos Generales')
            
            # Hoja de análisis con estilos
            if not df_pedidos.empty or not df_creditos.empty:
                analisis_data = self._generar_datos_analisis(df_pedidos, df_creditos)
                df_analisis = pd.DataFrame(analisis_data)
                self.crear_hoja_con_estilo(writer, df_analisis, 'Análisis')
        
        # Reabrir el archivo para aplicar estilos
        from openpyxl import load_workbook
        workbook = load_workbook(ruta_archivo)
        self.aplicar_estilos_excel(workbook)
        workbook.save(ruta_archivo)
        
        return ruta_archivo
    
    def _generar_datos_analisis(self, df_pedidos, df_creditos):
        """Genera los datos de análisis en formato más estructurado."""
        analisis = []
        
        if not df_pedidos.empty:
            analisis.extend([
                {
                    'Categoría': 'PEDIDOS',
                    'Métrica': 'Total de Pedidos',
                    'Valor': len(df_pedidos),
                    'Tipo': 'Cantidad',
                    'Descripción': 'Número total de pedidos realizados'
                },
                {
                    'Categoría': 'PEDIDOS',
                    'Métrica': 'Promedio por Pedido',
                    'Valor': round(df_pedidos['Total'].mean(), 2),
                    'Tipo': 'Monto',
                    'Descripción': 'Valor promedio de cada pedido'
                },
                {
                    'Categoría': 'PEDIDOS',
                    'Métrica': 'Pedido más Alto',
                    'Valor': df_pedidos['Total'].max(),
                    'Tipo': 'Monto',
                    'Descripción': 'Valor del pedido más costoso'
                },
                {
                    'Categoría': 'PEDIDOS',
                    'Métrica': 'Pedido más Bajo',
                    'Valor': df_pedidos['Total'].min(),
                    'Tipo': 'Monto',
                    'Descripción': 'Valor del pedido más económico'
                }
            ])
        
        if not df_creditos.empty:
            creditos_positivos = df_creditos[df_creditos['Status'] == 'Positivo']
            creditos_negativos = df_creditos[df_creditos['Status'] == 'Negativo']
            
            analisis.extend([
                {
                    'Categoría': 'CRÉDITOS',
                    'Métrica': 'Créditos Positivos',
                    'Valor': len(creditos_positivos),
                    'Tipo': 'Cantidad',
                    'Descripción': 'Número de créditos a favor'
                },
                {
                    'Categoría': 'CRÉDITOS',
                    'Métrica': 'Créditos Negativos',
                    'Valor': len(creditos_negativos),
                    'Tipo': 'Cantidad',
                    'Descripción': 'Número de deudas registradas'
                }
            ])
            
            if len(creditos_positivos) > 0:
                analisis.append({
                    'Categoría': 'CRÉDITOS',
                    'Métrica': 'Promedio Créditos Positivos',
                    'Valor': round(creditos_positivos['Monto'].mean(), 2),
                    'Tipo': 'Monto',
                    'Descripción': 'Valor promedio de créditos a favor'
                })
            
            if len(creditos_negativos) > 0:
                analisis.append({
                    'Categoría': 'CRÉDITOS',
                    'Métrica': 'Promedio Deudas',
                    'Valor': round(abs(creditos_negativos['Monto'].mean()), 2),
                    'Tipo': 'Monto',
                    'Descripción': 'Valor promedio de las deudas'
                })
        
        return analisis


def generar_reporte_gastos_diarios(fecha_reporte=None, nombre_archivo=None):
    """
    Función de conveniencia para generar un reporte de gastos diarios.
    """
    reporte = ReporteGastosDiarios(fecha_reporte)
    return reporte.generar_reporte_excel(nombre_archivo)


def generar_reporte_rango_fechas(fecha_inicio, fecha_fin, nombre_archivo=None):
    """
    Genera un reporte para un rango de fechas.
    
    Args:
        fecha_inicio (date): Fecha de inicio del rango.
        fecha_fin (date): Fecha de fin del rango.
        nombre_archivo (str): Nombre del archivo Excel.
    
    Returns:
        str: Ruta del archivo generado.
    """
    if not nombre_archivo:
        inicio_str = fecha_inicio.strftime("%Y%m%d")
        fin_str = fecha_fin.strftime("%Y%m%d")
        nombre_archivo = f"reporte_gastos_{inicio_str}_a_{fin_str}.xlsx"
    
    # Crear directorio de reportes si no existe
    directorio_reportes = "reportes"
    if not os.path.exists(directorio_reportes):
        os.makedirs(directorio_reportes)
    
    ruta_archivo = os.path.join(directorio_reportes, nombre_archivo)
    
    # Obtener datos del rango de fechas
    pedidos = Pedido.objects.filter(
        fecha__range=[fecha_inicio, fecha_fin]
    ).select_related('platillo', 'alumnoId', 'profesorId', 'nivelEducativo', 'encargadoId')
    
    creditos_diarios = CreditoDiario.objects.filter(
        fecha__range=[fecha_inicio, fecha_fin]
    ).select_related('tutorId', 'profesorId', 'pedido')
    
    creditos_generales = Credito.objects.filter(
        fecha__range=[fecha_inicio, fecha_fin]
    ).select_related('tutorId', 'profesorId')
    
    # Convertir pedidos a DataFrame
    data_pedidos = []
    for pedido in pedidos:
        data_pedidos.append({
            'ID Pedido': pedido.id,
            'Platillo': pedido.platillo.nombre,
            'Total': float(pedido.total),
            'Estado': pedido.get_status_label(),
            'Turno': pedido.get_turno_label(),
            'Alumno': str(pedido.alumnoId) if pedido.alumnoId else '',
            'Nivel Educativo': f"{getChoiceLabel(NIVELEDUCATIVO,pedido.nivelEducativo.nivel)} - {getChoiceLabel(GRADO,pedido.nivelEducativo.grado)} - {getChoiceLabel(GRUPO,pedido.nivelEducativo.grupo)}" if pedido.nivelEducativo else '',
            'Profesor': str(pedido.profesorId) if pedido.profesorId else '',
            'Fecha': pedido.fecha
        })
    
    # Convertir créditos diarios a DataFrame
    data_creditos_diarios = []
    for credito in creditos_diarios:
        status_credito = "Positivo" if credito.monto > 0 else "Negativo" if credito.monto < 0 else "Neutro"
        data_creditos_diarios.append({
            'ID Crédito': credito.id,
            'Monto': float(credito.monto),
            'Status': status_credito,
            'Tipo': 'Crédito Diario',
            'Tutor': str(credito.tutorId) if credito.tutorId else '',
            'Profesor': str(credito.profesorId) if credito.profesorId else '',
            'Fecha': credito.fecha
        })
    
    # Convertir créditos generales a DataFrame
    data_creditos_generales = []
    for credito in creditos_generales:
        status_credito = "Positivo" if credito.monto > 0 else "Negativo" if credito.monto < 0 else "Neutro"
        data_creditos_generales.append({
            'ID Crédito': credito.id,
            'Monto': float(credito.monto),
            'Status': status_credito,
            'Tipo': 'Crédito General',
            'Tutor': str(credito.tutorId) if credito.tutorId else '',
            'Profesor': str(credito.profesorId) if credito.profesorId else '',
            'Fecha': credito.fecha
        })
    
    df_pedidos = pd.DataFrame(data_pedidos)
    df_creditos_diarios = pd.DataFrame(data_creditos_diarios)
    df_creditos_generales = pd.DataFrame(data_creditos_generales)
    
    # Combinar créditos
    if not df_creditos_diarios.empty and not df_creditos_generales.empty:
        df_creditos_todos = pd.concat([df_creditos_diarios, df_creditos_generales], ignore_index=True)
    elif not df_creditos_diarios.empty:
        df_creditos_todos = df_creditos_diarios
    elif not df_creditos_generales.empty:
        df_creditos_todos = df_creditos_generales
    else:
        df_creditos_todos = pd.DataFrame()
    
    # Crear resumen del rango
    total_pedidos = df_pedidos['Total'].sum() if not df_pedidos.empty else 0
    total_creditos = df_creditos_todos['Monto'].sum() if not df_creditos_todos.empty else 0
    
    if not df_creditos_todos.empty:
        creditos_positivos = df_creditos_todos[df_creditos_todos['Status'] == 'Positivo']['Monto'].sum()
        creditos_negativos = df_creditos_todos[df_creditos_todos['Status'] == 'Negativo']['Monto'].sum()
    else:
        creditos_positivos = 0
        creditos_negativos = 0
    
    resumen_rango = {
        'Fecha Inicio': fecha_inicio,
        'Fecha Fin': fecha_fin,
        'Días en el Rango': (fecha_fin - fecha_inicio).days + 1,
        'Total Pedidos': float(total_pedidos),
        'Total Créditos': float(total_creditos),
        'Créditos Positivos': float(creditos_positivos),
        'Créditos Negativos': float(creditos_negativos),
        'Deuda Total Cafetería': float(abs(creditos_negativos)),
        'Balance Neto': float(total_creditos),
        'Cantidad Pedidos': len(df_pedidos) if not df_pedidos.empty else 0,
        'Cantidad Movimientos Crédito': len(df_creditos_todos) if not df_creditos_todos.empty else 0
    }
    
    df_resumen_rango = pd.DataFrame([resumen_rango])
    
    # Escribir a Excel
    with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
        # Hoja de resumen del rango
        df_resumen_rango.to_excel(writer, sheet_name='RESUMEN RANGO', index=False)
        
        # Hoja de pedidos del rango
        if not df_pedidos.empty:
            df_pedidos.to_excel(writer, sheet_name='Pedidos Rango', index=False)
        else:
            pd.DataFrame({'Mensaje': ['No hay pedidos en este rango de fechas']}).to_excel(
                writer, sheet_name='Pedidos Rango', index=False
            )
        
        # Hoja de todos los créditos del rango
        if not df_creditos_todos.empty:
            df_creditos_todos.to_excel(writer, sheet_name='Créditos Rango', index=False)
        else:
            pd.DataFrame({'Mensaje': ['No hay movimientos de crédito en este rango de fechas']}).to_excel(
                writer, sheet_name='Créditos Rango', index=False
            )
        
        # Hojas separadas para cada tipo de crédito
        if not df_creditos_diarios.empty:
            df_creditos_diarios.to_excel(writer, sheet_name='Créditos Diarios Rango', index=False)
        
        if not df_creditos_generales.empty:
            df_creditos_generales.to_excel(writer, sheet_name='Créditos Generales Rango', index=False)
    
    return ruta_archivo